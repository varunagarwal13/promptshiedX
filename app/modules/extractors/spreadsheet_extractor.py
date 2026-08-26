"""
PromptShield X - Spreadsheet Extractor
Parses Excel (.xlsx/.xls) and CSV files, detects DDE/Formula injection and hidden sheet payloads.
"""

import io
import csv
import openpyxl
from typing import Optional
from app.modules.extractors import BaseExtractor, ExtractedSegment, ExtractionResult, SourceType, ThreatCategory


class SpreadsheetExtractor(BaseExtractor):
    DDE_PREFIXES = ("=", "@", "+", "-", "|", "\t", "\r")

    def extract(self, data: bytes, filename: Optional[str] = None, **kwargs) -> ExtractionResult:
        result = ExtractionResult(source_type=SourceType.SPREADSHEET, filename=filename)
        is_csv = filename.lower().endswith(".csv") if filename else False

        if is_csv:
            self._extract_csv(data, result)
        else:
            self._extract_excel(data, result)

        result.raw_character_count = sum(len(s.content) for s in result.segments)
        result.anomalies_detected = sum(1 for s in result.segments if s.is_hidden or s.threat_indicators)
        return result

    def _extract_csv(self, data: bytes, result: ExtractionResult):
        try:
            text = data.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text))
            for row_idx, row in enumerate(reader, start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    val = cell_value.strip()
                    if not val:
                        continue
                    
                    threats = []
                    penalty = 0.0
                    if val.startswith(self.DDE_PREFIXES):
                        threats.append(ThreatCategory.FORMULA_INJECTION)
                        penalty += 35.0

                    result.segments.append(
                        ExtractedSegment(
                            content=val,
                            source_type=SourceType.SPREADSHEET,
                            location=f"CSV:R{row_idx}C{col_idx}",
                            is_hidden=False,
                            threat_indicators=threats,
                            confidence_penalty=penalty
                        )
                    )
        except Exception as e:
            result.extraction_warnings.append(f"CSV parsing error: {str(e)}")

    def _extract_excel(self, data: bytes, result: ExtractionResult):
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=False, read_only=False)
            
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                is_sheet_hidden = ws.sheet_state != "visible"
                
                for row in ws.iter_rows(values_only=False):
                    for cell in row:
                        val = cell.value
                        if val is None:
                            continue
                        
                        str_val = str(val).strip()
                        if not str_val:
                            continue

                        is_hidden = is_sheet_hidden
                        threats = []
                        penalty = 0.0

                        # Check if sheet is hidden
                        if is_sheet_hidden:
                            threats.append(ThreatCategory.HIDDEN_TEXT)
                            penalty += 30.0

                        # Check if row or column is hidden
                        if ws.row_dimensions[cell.row].hidden or (
                            cell.column_letter in ws.column_dimensions and ws.column_dimensions[cell.column_letter].hidden
                        ):
                            is_hidden = True
                            threats.append(ThreatCategory.HIDDEN_TEXT)
                            penalty += 30.0

                        # Check Formula / DDE injection
                        if str_val.startswith(self.DDE_PREFIXES) or cell.data_type == 'f':
                            threats.append(ThreatCategory.FORMULA_INJECTION)
                            penalty += 40.0

                        # Check Camouflaged Font (Font color == Cell Background color)
                        if cell.fill and cell.fill.start_color and cell.font and cell.font.color:
                            if cell.fill.start_color.rgb == cell.font.color.rgb:
                                is_hidden = True
                                threats.append(ThreatCategory.HIDDEN_TEXT)
                                penalty += 50.0

                        result.segments.append(
                            ExtractedSegment(
                                content=str_val,
                                source_type=SourceType.SPREADSHEET,
                                location=f"{sheetname}!{cell.coordinate}",
                                is_hidden=is_hidden,
                                threat_indicators=threats,
                                confidence_penalty=penalty,
                                metadata={
                                    "sheet": sheetname,
                                    "cell": cell.coordinate,
                                    "data_type": cell.data_type
                                }
                            )
                        )
            wb.close()
        except Exception as e:
            result.extraction_warnings.append(f"Excel parsing error: {str(e)}")